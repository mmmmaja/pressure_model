# node (sensor) coordinates ventral (plus pressure dummy data)
import os

import numpy as np
import openpyxl
from matplotlib import pyplot as plt

"""
In this file I am reading the data from robotic arm recordings. The data is stored in .csv files.
"""


nds_frontX = [0.000, 25.981, 51.962, 77.943, 12.990, 38.971, 64.952, 90.933, 0.000, 25.981, 51.962, 77.943, 12.990,
              38.971, 64.952, 90.933, 0.000, 25.981, 51.962, 77.943, 12.990, 38.971, 64.952, 90.933, 0.000, 25.981,
              51.962, 77.943, 12.990, 38.971, 64.952, 90.933, 0.000, 25.981, 51.962, 77.943, 12.990, 38.971, 64.952,
              90.933, 0.000, 25.981, 51.962, 77.943, 12.990, 38.971, 64.952, 90.933]
nds_frontY = [0.000, 0.000, 0.000, 0.000, 7.500, 7.500, 7.500, 7.500, 15.000, 15.000, 15.000, 15.000, 22.500, 22.500,
              22.500, 22.500, 30.000, 30.000, 30.000, 30.000, 37.500, 37.500, 37.500, 37.500, 45.000, 45.000, 45.000,
              45.000, 52.500, 52.500, 52.500, 52.500, 60.000, 60.000, 60.000, 60.000, 67.500, 67.500, 67.500, 67.500,
              75.000, 75.000, 75.000, 75.000, 82.500, 82.500, 82.500, 82.500]
nds_frontZ = [5.50, 2.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.200, 0.000, 0.000, 0.000, 0.000,
              0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000,
              0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000,
              0.000, 0.000, 0.000]
pressures = [10.000, 15.000, 0.000, 0.000, 5.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000,
             0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000,
             0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 0.000,
             0.000, 0.000, 0.000]

# node (sensor) coordinates dorsal top
nds_backTopX = [0.000, 25.981, 51.962, 77.943, 103.924, 129.905, 12.990, 38.971, 64.952, 90.933, 116.915, 0.000, 25.981,
                51.962, 77.943, 103.924, 129.905, 12.990, 38.971, 64.952, 90.933, 116.915]
nds_backTopY = [0.000, 0.000, 0.000, 0.000, 0.000, 0.000, 7.500, 7.500, 7.500, 7.500, 7.500, 15.000, 15.000, 15.000,
                15.000, 15.000, 15.000, 22.500, 22.500, 22.500, 22.500, 22.500]

# node (sensor) coordinates dorsal bottom
nds_backBottomX = [0.000, 25.981, -12.990, 12.990, 38.971, 0.000, 25.981, -12.990, 12.990, 38.971]
nds_backBottomY = [0.000, 0.000, 7.500, 7.500, 7.500, 15.000, 15.000, 22.500, 22.500, 22.500]


def get_sensor_positions():
    sensor_coords = []

    for i in range(len(nds_frontX)):
        sensor_coords.append([
            nds_frontX[i], nds_frontY[i], 0.0
        ])
    return np.array(sensor_coords)


def plot_sensor_positions():
    # Create a 3D plot
    fig = plt.figure()
    ax = fig.add_subplot(projection='3d')

    sensor_positions = get_sensor_positions()
    x = sensor_positions[:, 0]
    y = sensor_positions[:, 1]
    z = sensor_positions[:, 2]

    # Plot the surface
    ax.scatter(x, y, z)

    plt.show()


def read_lukas_recording(path):
    sheet = openpyxl.load_workbook(path).active

    time = read_time(sheet)
    sensor_positions = get_sensor_positions()
    sensor_reading = read_sensor_data(sheet, path)

    return sensor_positions, sensor_reading, time


def read_time(sheet):
    """
    :return: time data from Excel file
    """
    num_measurements = sheet.max_column
    time = np.zeros(num_measurements)
    for i in range(1, num_measurements):
        # Read time in secs
        time[i] = sheet.cell(column=i, row=2).value
    return time


def read_sensor_data(sheet, path):
    # shape=(num_sensors, num_measurements)
    pressure = read_pressure_front(sheet, path)
    return pressure.T


def read_pressure_front(sheet, path):
    """
    :return: Pressures measurements from front patch
    """
    num_measurements = sheet.max_column
    num_sensors = sheet.max_row - 6 if path[-8:-5] == "LIN" else sheet.max_row - 3

    pressure = np.zeros(shape=(48, num_measurements))
    row_count = 0
    for row in sheet.iter_rows(
            min_row=4,
            min_col=1,
            max_row=min(num_sensors, 48) + 3,
            max_col=num_measurements,
            values_only=True
    ):
        column_count = 0
        for value in row:
            value = max(float(value), 0.0)
            pressure[row_count, column_count] = value
            column_count += 1
        row_count += 1
    return pressure


def read_pressure_back_top(sheet, path):
    """
    :return: Pressures measurements from back top patch
    """
    num_measurements = sheet.max_column
    num_sensors = sheet.max_row - 6 if path[-8:-5] == "LIN" else sheet.max_row - 3

    pressure = np.zeros(shape=(22, num_measurements))
    # Read pressure values for the top back array
    row_count = 0
    for row in sheet.iter_rows(
            min_row=49 + 3,
            min_col=1,
            max_row=min(num_sensors, 70) + 3,
            max_col=num_measurements,
            values_only=True
    ):
        column_count = 0
        for value in row:
            value = max(float(value), 0.0)
            pressure[row_count, column_count] = value
            column_count += 1
        row_count += 1
    return pressure


def read_pressure_back_bottom(sheet, path):
    num_measurements = sheet.max_column
    num_sensors = sheet.max_row - 6 if path[-8:-5] == "LIN" else sheet.max_row - 3

    pressure = np.zeros(shape=(10, num_measurements))
    row_count = 0
    for row in sheet.iter_rows(
            min_row=71 + 3,
            min_col=1,
            max_row=num_sensors + 3,
            max_col=num_measurements,
            values_only=True
    ):
        column_count = 0
        for value in row:
            value = max(float(value), 0.0)
            pressure[row_count, column_count] = value
            column_count += 1
        row_count += 1
    return pressure
